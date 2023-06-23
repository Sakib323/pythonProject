from selenium import webdriver;
import os
import time
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import pandas as pd
import openpyxl

filename = "existing_file.xlsx"
if not os.path.isfile(filename):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    headers = ["Permit", 'Date',"Entry Date","Issue Date","Permit Type","Permit Status",'Project #', 'Type of Building', 'Project Name', 'Equipment Type', 'Property Use',
               'Contract Cost',"Electrical_Construction_Costs","Total_Electrical_Cost","Total_Equipment_Fee","Building_Contract_Cost","Sub Permits Contract Cost",'USDC Code', 'Owner_Tenant', 'Owner_Tenant_Phone', 'Owner_Tenant_Address',
               'Contractor_Name', 'Contractor_ID', 'Contractor_Email_Address', 'Contractor_Phone', 'Contractor_Address',
               'License']
    for col_num, header_value in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_num).value = header_value
    workbook.save(filename)
else:
    workbook = openpyxl.load_workbook(filename)


def write(permit, Date,Entry_Date,Issue_Date,Permit_Type,Permit_Status, Project, Type_of_Building, Project_Name, Equipment_Type, Property_Use,
          Contract_Cost,Electrical_Construction_Costs,Total_Electrical_Cost, Total_Equipment_Fee,Building_Contract_Cost,Sub_Permits_Contract_Cost,USDC_Code, Owner_Tenant, Owner_Tenant_Phone, Owner_Tenant_Address,
          Contractor_Name, Contractor_ID, Contractor_Email_Address, Contractor_Phone, Contractor_Address, License):
    sheet = workbook.active
    new_data = [[permit, Date,Entry_Date,Issue_Date,Permit_Type,Permit_Status,Project, Type_of_Building, Project_Name, Equipment_Type, Property_Use,
                 Contract_Cost,Electrical_Construction_Costs,Total_Electrical_Cost, Total_Equipment_Fee,Building_Contract_Cost,Sub_Permits_Contract_Cost,USDC_Code, Owner_Tenant, Owner_Tenant_Phone, Owner_Tenant_Address,
                 Contractor_Name, Contractor_ID, Contractor_Email_Address, Contractor_Phone, Contractor_Address,
                 License]]
    next_row = sheet.max_row + 1
    for row_data in new_data:
        for col_num, cell_value in enumerate(row_data, start=1):
            sheet.cell(row=next_row, column=col_num).value = cell_value
        next_row += 1
    workbook.save(filename)


by_new_line = []
phone = []
address = []
Contractor = []
Date = []
Contract_Cost_ = []
Type_of_Building = ""
License = ""
Owner_or_Tenant_name = ""
Entered_Date = ""
project = ""
Issued_Date = ""
Contractor_Name = ""
Contractor_ID = ""
Owner_or_Tenant_phone = ""
Contractor_Phone = ""
Owner_or_Tenant_address = ""
Contractor_address = ""
Contractor_Email_Address = ""
Project_Name = ""
Equipment_Type = ""
Property_Use = ""
Contract_Cost = ""
USDC_Code = ""
index_of_lic = ""
index_of_owner_phone = ""
index_of_contractor_phone = ""
Permit_Type=""
Permit_Status=""
Electrical_Construction_Costs=""
Total_Electrical_Cost=""
Total_Equipment_Fee=""
Building_Contract_Cost=""
Sub_Permits_Contract_Cost=""
index_of_con_add=""

df = pd.read_excel("permit.xlsx")
permit = df['Permit#']

sleep_tune=2

driver = webdriver.Chrome();

try:
    driver.get("https://webpermit.mecklenburgcountync.gov/Default.aspx?PossePresentation=SearchByPermit")
except:
    print("need a better and stable internet connection and make sure you have the chrome window open not minimized or polluated with other window")

for permit_nmbr in permit:


    try:
        time.sleep(sleep_tune)
        input_field = driver.find_element(By.ID, "ExternalFileNum_467186_S0")
        input_field.clear()
        input_field.send_keys(permit_nmbr)
        input_field.send_keys(Keys.ENTER)
        time.sleep(sleep_tune)
    except:
        print("need a better and stable internet connection and make sure you have the chrome window open not minimized or polluated with other window")

    try:
        all_data = driver.find_element(By.CLASS_NAME, "possedetail").text

        for data in all_data.split("\n"):

            by_new_line.append(data)

            for data in by_new_line:
                if (data.__contains__("Owner/Tenant")):
                    index_of_owner_name = data
                    Owner_or_Tenant_name = data.replace("Owner/Tenant", "")
                if (data.__contains__("Phone")):
                    for info in data.split("\n"):
                        phone.append(info)

                if (data.__contains__("Email Address")):
                    for info in data.split("\n"):
                        Contractor_Email_Address = info.replace("Email Address", "")
                        index_of_contractor_phone = info

                if (data.__contains__("Address")):

                    for info in data.split("\n"):
                        address.append(info)
                if (data.__contains__("Contractor")):
                    for info in data.split("\n"):
                        Contractor.append(info)
                if (data.__contains__("Entry Date")):
                    for info in data.split("\n"):
                        Date.append(info)

                if (data.__contains__("Project #")):
                    for info in data.split("\n"):
                        project = info.replace("Project #", "")
                if (data.__contains__("Type of Building")):
                    for info in data.split("\n"):
                        Type_of_Building = info.replace("Type of Building", "").replace("Project #", "").replace(
                            project, "").replace("Equipment Changeout", "")
                if (data.__contains__("License #")):
                    for info in data.split("\n"):
                        License = info.replace("License #", "")
                        index_of_lic = info

                if (data.__contains__("Project Name")):
                    for info in data.split("\n"):
                        Project_Name = info.replace("Project Name", "").replace("Equipment Type", "")
                if (data.__contains__("Equipment Type")):
                    for info in data.split("\n"):
                        Equipment_Type = info.replace("Equipment Type", "").replace("Project Name", "")
                if (data.__contains__("Property Use")):
                    for info in data.split("\n"):
                        Property_Use = info.replace("Property Use", "")

                if (data.__contains__("Contract Cost")):
                    for info in data.split("\n"):
                        Contract_Cost_.append(info)
                if (data.__contains__("USDC Code")):
                    for info in data.split("\n"):
                        USDC_Code = info.replace("USDC Code", "")
                if (data.__contains__("Contractor ID")):
                    for info in data.split("\n"):
                        Contractor_ID = info.replace("Contractor ID", "")
                if (data.__contains__("Permit Type")):
                    for info in data.split("\n"):
                        Permit_Type = info.replace("Permit Type", "")
                if (data.__contains__("Permit Status")):
                    for info in data.split("\n"):
                        Permit_Status = info.replace("Permit Status", "")

                if (data.__contains__("Electrical Construction Costs")):
                    for info in data.split("\n"):
                        Electrical_Construction_Costs = info.replace("Electrical Construction Costs", "")

                if (data.__contains__("Total Electrical Cost")):
                    for info in data.split("\n"):
                        Total_Electrical_Cost = info.replace("Total Electrical Cost", "")
                if (data.__contains__("Total Equipment Fee")):
                    for info in data.split("\n"):
                        Total_Equipment_Fee = info.replace("Total Equipment Fee", "")
                if (data.__contains__("Building Contract Cost")):
                    for info in data.split("\n"):
                        Building_Contract_Cost = info.replace("Building Contract Cost", "")

                if (data.__contains__("Sub Permits Contract Cost")):
                    for info in data.split("\n"):
                        Sub_Permits_Contract_Cost = info.replace("Sub Permits Contract Cost", "")

        try:
            index_of_con_add = abs((by_new_line.index(index_of_contractor_phone) - by_new_line.index(index_of_lic)))
            index_nmbr = []
            add = " "
            for i in range(index_of_con_add):
                if (i != 0):
                    if (by_new_line.index(index_of_contractor_phone) > by_new_line.index(index_of_lic)):
                        index_nmbr.append(by_new_line.index(index_of_lic) + i)
                    else:
                        index_nmbr.append(by_new_line.index(index_of_contractor_phone) + i)

            for index in index_nmbr:
                add = add + by_new_line[index]

            # print(index_nmbr)
            add = add.replace(by_new_line[index_nmbr[0]], "")

            if (add.__contains__("Address")):
                add = add.replace("Address", "")


        except:
            print("Error here")

        try:
            Permit_Type = Permit_Type.replace("Permit #", "").replace(permit_nmbr, "").replace("Inspection Results", "")
        except:
            print("Error here")

        try:
            Permit_Status = Permit_Status.split()[len(Permit_Status.split()) - 1]
        except:
            print("Error here")

        try:
            Entered_Date = Date[0].replace("Entry Date", "")
        except:
            print("Error here")

        try:
            Contractor_Name = Contractor[0].replace("Contractor", "")
        except:
            print("Error here")

        try:
            Owner_or_Tenant_phone = phone[0].replace("Phone", "").replace(" ", "")
        except:
            print("Error here")

        try:
            Contractor_Phone = by_new_line[by_new_line.index(index_of_contractor_phone) + 1].replace("Phone", "")
        except:
            print("Error here")
        try:
            Owner_or_Tenant_address = by_new_line[by_new_line.index(phone[0]) + 1].replace("Address", "")
        except:
            print("Error here")

        try:
            Equipment_Type = Equipment_Type.split()[len(Equipment_Type.split()) - 1]
        except:
            print("Error here")

        try:
            Contractor_address = by_new_line[by_new_line.index(index_of_lic) - 1]
        except:
            print("Error here")

        try:
            Project_Name = Project_Name.replace(Equipment_Type, "")
        except:
            print("Error here")

        try:
            Contract_Cost = Contract_Cost_[0].replace("Contract Cost", "").replace(Property_Use, "").replace(
                "Property Use", "")
        except:
            print("Error here")

        try:
            Contract_Cost = Contract_Cost.split()[len(Contract_Cost.split()) - 1]
        except:
            print("Error here")

        try:
            Property_Use = Property_Use.replace(Contract_Cost, "").replace("Contract Cost", "")
        except:
            print("Error here")

        try:
            Contractor_address = by_new_line[by_new_line.index(index_of_lic) - 1]
        except:
            print("Error here")

        Electrical_Construction_Costs = Electrical_Construction_Costs.replace(":", "")
        Total_Electrical_Cost = Total_Electrical_Cost.replace(":", "")
        Total_Equipment_Fee = Total_Equipment_Fee.replace(":", "")
        Building_Contract_Cost = Building_Contract_Cost.replace(":", "")
        Sub_Permits_Contract_Cost = Sub_Permits_Contract_Cost.replace(":", "")

        project = project.replace("Type of Building", "").split()[0]
        Type_of_Building = Type_of_Building.replace(project, "")
        Entered_Date = Entered_Date.split()[0]

        if (Permit_Status.__contains__("Issued")):
            Issued_Date = Entered_Date

        write(permit_nmbr, Entered_Date, Entered_Date, Issued_Date, Permit_Type, Permit_Status, project,
              Type_of_Building, Project_Name, Equipment_Type, Property_Use,
              Contract_Cost, Electrical_Construction_Costs, Total_Electrical_Cost, Total_Equipment_Fee,
              Building_Contract_Cost, Sub_Permits_Contract_Cost, USDC_Code, Owner_or_Tenant_name, Owner_or_Tenant_phone,
              Owner_or_Tenant_address,
              Contractor_Name, Contractor_ID, Contractor_Email_Address, Contractor_Phone, add, License)

        by_new_line = []
        phone = []
        address = []
        Contractor = []
        Date = []
        Contract_Cost_ = []
        Type_of_Building = ""
        License = ""
        Owner_or_Tenant_name = ""
        Entered_Date = ""
        project = ""
        Issued_Date = ""
        Contractor_Name = ""
        Contractor_ID = ""
        Owner_or_Tenant_phone = ""
        Contractor_Phone = ""
        Owner_or_Tenant_address = ""
        Contractor_address = ""
        Contractor_Email_Address = ""
        Project_Name = ""
        Equipment_Type = ""
        Property_Use = ""
        Contract_Cost = ""
        USDC_Code = ""
        index_of_lic = ""
        index_of_owner_phone = ""
        index_of_contractor_phone = ""
        Permit_Type = ""
        Permit_Status = ""
        Electrical_Construction_Costs = ""
        Total_Electrical_Cost = ""
        Total_Equipment_Fee = ""
        Building_Contract_Cost = ""
        Sub_Permits_Contract_Cost = ""
        index_of_con_add = ""
    except:
        print("No permit found by number: "+permit_nmbr)

    driver.execute_script("window.history.go(-1)")

driver.quit()

